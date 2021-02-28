"""
Exports a spill report to an Excel file for download.
Future: send email notifications
"""

import os
from datetime import datetime
from glob import glob
import traceback
from openpyxl import Workbook, load_workbook

from app.reports import reports_db
import settings

logger = settings.setup_logger(__name__)
exports = settings.report_exports


def list_report_fields():
    sr = reports_db.SpillReport()
    # List all fields for export
    field_names = []
    for name in sr.__dir__():
        if not name.startswith('_'):
            field_names.append(name)
    if 'metadata' in field_names:
        field_names.remove('metadata')
    return field_names


def cleanup_old_reports():
    # Delete any existing report
    exist_reports = glob(os.path.join(exports, 'report*.xlsx'))
    for existing in exist_reports:
        logger.warning('Delete old report dump: %s' % existing)
        try:
            os.remove(existing)
        except:
            logger.error(traceback.format_exc())
            pass


def dump_all_excel():
    cleanup_old_reports()

    book = Workbook()
    sheet = book.active

    field_names = list_report_fields()
    sheet.append(field_names)

    session = reports_db.Session()
    results = session.query(reports_db.SpillReport).all()
    session.close()
    logger.info(field_names)
    logger.info('Dumping %s report to excel' % len(results))
    for res in results:
        report = reports_db.result_to_dict(res)
        row_data = []
        for field in field_names:
            val = report.get(field)
            if type(val) is datetime:
                val = val.strftime(settings.display_date_fmt)
            row_data.append(val)
        logger.info(row_data)
        sheet.append(row_data)

    fname = 'reports_%s.xlsx' % datetime.now().strftime(settings.filesafe_timestamp)
    export_file = os.path.join(exports, fname)
    logger.info('Save data to %s' % export_file)
    book.save(export_file)
    book.close()

    return export_file


def named_range_to_cell(book, name):
    # Assume that the named range refers to only one cell, not a whole range
    try:
        range_obj = book.defined_names[name]
        for title, coord in range_obj.destinations:
            return coord
    except KeyError:
        return None


def report_to_excel(report):
    cleanup_old_reports()
    # Set the output filename
    report_num = report.get('report_num')
    last_updated = report.get('last_updated')
    export_file = os.path.join(exports, 'report_%s_v%s.xlsx' % (
        report_num, last_updated.strftime(settings.filesafe_timestamp)))
    logger.info('Export to: %s' % export_file)
    spill_template = os.path.join(settings.templates, 'SpillTemplate.xlsx')

    # Open spill template, add content using named ranges
    book = load_workbook(spill_template)
    sheet = book.active

    field_names = list_report_fields()
    fcount = 0
    for name in field_names:
        # Skip if nothing in this field, or no matching named range in Excel template
        val = report.get(name)
        if not val:
            continue
        # Convert datetimes to display format
        if type(val) is datetime:
            val = val.strftime(settings.display_date_fmt)
        coord = named_range_to_cell(book, name)
        if not coord:
            logger.warning('Excel template is missing the field: %s' % name)
            continue
        sheet[coord] = val
        fcount += 1
    logger.info('Set values for %s fields' % fcount)
    book.save(export_file)
    book.close()

    return export_file


def main():
    dump_all_excel()


if __name__ == '__main__':
    main()

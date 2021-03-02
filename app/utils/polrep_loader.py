from openpyxl import load_workbook
import os
import settings
from datetime import datetime
from traceback import format_exc
from sqlalchemy import create_engine

engine = create_engine(settings.SPILL_TRACKER_DB_URL)

logger = settings.setup_logger(__name__)

from app.reports.reports_db import db, SpillReport
from app.app_factory import create_app

logger.info('Creating app object in %s' % __name__)

app = create_app()


def open_legacy_file(path):
    book = load_workbook(path)
    return book


def fix_2017_format(report_num):
    rep_num = report_num.replace('-2017', '')
    return '2017-%s' % rep_num


def erase_all_reports():
    with app.app_context():
        try:
            logger.warning('About to delete all Spill Reports!')
            num_rows_deleted = db.session.query(SpillReport).delete()
            db.session.commit()
            logger.warning('%s reports deleted.' % num_rows_deleted)
        except:
            logger.error(format_exc())
            db.session.rollback()


# Create a full table matching report_map_view
# This is only called once, after loading legacy POLREPs
def create_report_map_tbl():
    logger.info('Creating report_map TABLE in PostGIS')
    engine.execute('DROP TABLE IF EXISTS report_map_tbl;')
    engine.execute('CREATE TABLE report_map_tbl AS SELECT * FROM report_map_view;')
    # Add geo index to the geometry column
    add_geo_index = '''
    CREATE INDEX idx_geom_report_map
    ON report_map_tbl
    USING GIST (geometry);
    '''
    engine.execute(add_geo_index)


def load_data_all(legacy_file):

    erase_all_reports()

    logger.info('Opening: %s' % legacy_file)
    book = open_legacy_file(legacy_file)
    for year in ['2017', '2018', '2019']:
        load_data_year(year, book)
    logger.info('Finished loading all legacy reports')
    create_report_map_tbl()
    logger.info('Done with legacy POLREPs. Please check PostGIS table.')


def load_data_year(year, book):

    logger.info('POLREPS for %s...' % year)
    year_ws = book[year]
    for i, row in enumerate(year_ws.iter_rows(min_row=2)):  # , max_row=2
        report_num = row[0].value
        if report_num is None:
            # Probably end of the table
            logger.warning('Probably end of the table? but openpyxl goes over a few rows...')
            continue

        # Fix 2017 report format
        if year == '2017':
            report_num = fix_2017_format(report_num)

        report_name = row[1].value
        report_label = 'Report: %s, %s' % (report_num, report_name)
        # date comes in as a datetime object with 00:00 timestamp
        spill_date = row[2].value
        # time is (usually) a string, need to add to date
        time_str = row[3].value

        # Try to parse as datetime if not null, and not already a datetime
        TIME_SHORT = '%H:%M'
        TIME_LONG = '%H:%M:%S'
        TIME_AMPM = '%I:%M:%S %p'
        if time_str:
            # Default: assume time_str was parsed as datetime.time
            time = time_str
            if type(time_str) is str:

                time_str = time_str.replace(';', ':')

                time_format = TIME_SHORT
                if 'am' in time_str.lower() or 'pm' in time_str.lower():
                    time_format = TIME_AMPM
                if time_str.count(':') == 2:
                    time_format = TIME_LONG

                # Try to parse as datetime
                try:
                    time = datetime.strptime(time_str, time_format).time()
                except (ValueError, TypeError):
                    logger.error('%s: Very Bad timestamp: %s' % (
                        report_label, time_str))
                    time = None

        else:
            logger.warning('Time field is null/empty')
            time = None

        # Update date with hour and minute
        if type(spill_date) is datetime and time is not None:
            try:
                spill_date = spill_date.replace(hour=time.hour, minute=time.minute)
            except:
                logger.warning('Time is empty or invalid')

        # Use spill date for last updated, if it is a valid datetime
        if type(spill_date) is datetime:
            last_updated = spill_date
        else:
            logger.warning('Spill date is invalid, using current date.')
            last_updated = datetime.now()

        vessel_name = row[4].value
        pollutant_details = row[5].value
        pollutant = row[6].value

        # Get latitude, longitude, check type
        latitude = row[7].value  # float
        longitude = row[8].value  # float
        try:
            latitude = float(latitude)
        except:
            logger.error('%s: Invalid latitude: "%s" ' % (report_label, latitude))
            latitude = None

        try:
            longitude = float(longitude)
        except:
            logger.error('%s: Invalid longitude: "%s" ' % (report_label, longitude))
            longitude = None

        vessel_additional_info = row[9].value
        er_region = row[10].value
        response_activated = row[11].value
        fleet_tasking = row[12].value
        station_or_ship = row[13].value
        unit = row[14].value

        # The following needs an app context
        with app.app_context():
            # import flask_whooshalchemy as wa
            # logger.info('Adding whoosh fulltext index from app_context WITHIN polrep_loader')
            # wa.search_index(app, SpillReport)

            # Make a SpillReport object
            sr = SpillReport(
                # Add required fields
                last_updated=last_updated, recorded_by='CCG Legacy', user_id=0,
                report_num=report_num, report_name=report_name, spill_date=spill_date,
                vessel_name=vessel_name, pollutant_details=pollutant_details,
                pollutant=pollutant, latitude=latitude, longitude=longitude,
                vessel_additional_info=vessel_additional_info, er_region=er_region,
                response_activated=response_activated, fleet_tasking=fleet_tasking,
                station_or_ship=station_or_ship, unit=unit)

            try:
                db.session.add(sr)
                logger.info('Added to db session: %s %s %s %s' % (report_num, report_name, spill_date, time_str))
                db.session.commit()
                # if i % 100 == 0:
                #     db.session.commit()
                #     logger.info('Commit %s' % i)
            except:
                logger.error(format_exc())

    # Commit any remaining rows
    # with app.app_context():
    #     try:
    #         db.session.commit()
    #         logger.info('Committed all remaining reports.')
    #     except:
    #         logger.error(format_exc())

    logger.info('Finished loading %s' % year)


def main():
    # legacy = os.path.join(settings.upload_root, '2021-02-11_Legacy_Polreps.xlsx')
    # open_legacy_file(legacy)
    # load_data_year('2017', legacy)
    # load_data_all(legacy)
    logger.warning('DO NOT RUN LEGACY POLREP LOADER FROM HERE! Only use util_scripts.py')


if __name__ == '__main__':
    main()

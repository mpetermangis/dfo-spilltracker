from openpyxl import Workbook, load_workbook
import os
import settings

logger = settings.setup_logger(__name__)


def open_legacy_file(path):
    book = load_workbook(path)
    return book


def load_data(year, path):
    book = open_legacy_file(path)
    year = book[year]
    for row in year.iter_rows(min_row=2, max_row=2):
        report_num = row[0].value
        report_name = row[1].value
        # date comes in as a datetime object with 00:00 timestamp
        date = row[2].value
        # time is a string, need to add to date
        time = row[3].value
        latitude = row[7].value  # float
        longitude = row[8].value  # float
        logger.info('%s %s %s %s' % (report_num, report_name, date, time))
        # logger.info(report_num)

        # for cell in row:
        #     print(cell.value)


def main():
    legacy = os.path.join(settings.upload_root, '2021-02-11_Legacy_Polreps.xlsx')
    # open_legacy_file(legacy)
    load_data('2017', legacy)


if __name__ == '__main__':
    main()
